# utils/rt_export_manager.py
import os, time, json, threading
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def _flatten(obj, prefix="", out=None):
    if out is None:
        out = {}
    if not isinstance(obj, dict):
        return out
    for k, v in obj.items():
        key = f"{prefix}.{k}" if prefix else k
        if isinstance(v, dict):
            _flatten(v, key, out)
        else:
            out[key] = v
    return out

class RtExportManager:
    """
    Grabación RT a Excel:
    - start(tags): inicia nuevo xlsx (timestamp + tags)
    - ingest(sample): mete una fila (si active)
    - stop(): finaliza y deja listo para download
    - status(): estado + contador
    """
    def __init__(self, out_dir="exports", checkpoint_s=1.5):
        self.out_dir = out_dir
        self.checkpoint_s = checkpoint_s
        os.makedirs(out_dir, exist_ok=True)

        self._lock = threading.Lock()
        self.active = False
        self.tags = []
        self.rows_written = 0
        self.path = None
        self.started_at = None

        self._wb = None
        self._ws = None
        self._last_save = 0.0

    def start(self, tags: list[str]) -> dict:
        tags = [t for t in tags if isinstance(t, str) and t.strip()]
        if not tags:
            raise ValueError("tags vacío")

        with self._lock:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            self.path = os.path.join(self.out_dir, f"rt_export_{ts}.xlsx")

            self._wb = Workbook()
            self._ws = self._wb.active
            self._ws.title = "rt"

            self.tags = tags
            self.rows_written = 0
            self.started_at = time.time()
            self.active = True

            # header
            self._ws.append(["timestamp"] + self.tags)
            self._last_save = 0.0
            self._save(force=True)

            return self.status()

    def ingest(self, sample: dict):
        with self._lock:
            if not self.active or not self._ws:
                return

            if not isinstance(sample, dict):
                return

            flat = _flatten(sample)

            # timestamp robusto
            ts_raw = sample.get("timestamp", flat.get("timestamp", time.time()))

            try:
                if isinstance(ts_raw, (int, float)):
                    dt = datetime.fromtimestamp(float(ts_raw))
                elif isinstance(ts_raw, str):
                    s = ts_raw.strip()
                    # string numérico epoch
                    try:
                        dt = datetime.fromtimestamp(float(s))
                    except Exception:
                        # ISO string (ej. 2026-02-20T15:41:36.833Z)
                        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
                else:
                    dt = datetime.now()
            except Exception:
                dt = datetime.now()

            dt_str = dt.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]

            row = [dt_str] + [flat.get(t) for t in self.tags]
            self._ws.append(row)
            self.rows_written += 1

            # debug temporal
            if self.rows_written <= 3 or self.rows_written % 20 == 0:
                import logging
                logging.getLogger("uvicorn").info(
                    "RT export wrote row #%s | ts=%s",
                    self.rows_written, dt_str
                )

            self._save()

    def stop(self) -> dict:
        with self._lock:
            if not self.active:
                return self.status()
            self.active = False
            self._save(force=True)
            return self.status()

    def _save(self, force=False):
        if not self._wb or not self.path:
            return
        now = time.time()
        if force or (now - self._last_save) >= self.checkpoint_s:
            # ancho básico
            try:
                for i in range(1, 2 + len(self.tags)):
                    col = get_column_letter(i)
                    self._ws.column_dimensions[col].width = 24
            except Exception:
                pass

            self._wb.save(self.path)
            self._last_save = now

    def status(self) -> dict:
        return {
            "active": self.active,
            "rows_written": self.rows_written,
            "path": self.path,
            "tags": self.tags,
            "started_at": self.started_at,
        }
