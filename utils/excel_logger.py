# utils/excel_logger.py
import os, json, time, queue, threading
from datetime import datetime
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

def _flatten(obj, prefix="", out=None):
    if out is None: out = {}
    if not isinstance(obj, dict):
        return out
    for k, v in obj.items():
        key = f"{prefix}.{k}" if prefix else k
        if isinstance(v, dict):
            _flatten(v, key, out)
        else:
            out[key] = v
    return out

def _beautify(col: str) -> str:
    # "REAL.fast_1" -> "fast_1", "UDINT.tri_half" -> "tri_half"
    return col.split(".", 1)[-1] if "." in col else col

def _next_indexed_path(path_base: str) -> str:
    # si existe rt_2025-09-10.xlsx, crea rt_2025-09-10_02.xlsx, etc.
    if not os.path.exists(path_base):
        return path_base
    root, ext = os.path.splitext(path_base)
    idx = 2
    while True:
        cand = f"{root}_{idx:02d}{ext}"
        if not os.path.exists(cand):
            return cand
        idx += 1

class ExcelLogger(threading.Thread):
    """
    Escritor RT para Excel:
      • Hoja 'rt' primero: timestamp + columnas seleccionadas (cabeceras bonitas).
      • Hoja 'raw' segundo: timestamp + json crudo (por si acaso).
      • No guarda en disco en cada flush: checkpoint periódico.
    """
    def __init__(self, q: queue.Queue,
                 path_template: str = "exports/rt_{date}.xlsx",
                 flush_every: int = 50,          # junta N muestras
                 flush_interval: float = 0.2,    # o cada X s
                 sheet_name: str = "rt",
                 pretty_headers: bool = True,
                 long_format: bool = False,      # dejamos wide (cada tag es columna)
                 with_table: bool = True,
                 autosize: bool = True,
                 save_checkpoint_s: float = 2.0  # guarda a disco cada X s
                 ):
        super().__init__(daemon=True)
        self.q = q
        self.path_template = path_template
        self.flush_every = flush_every
        self.flush_interval = flush_interval
        self.sheet_name = sheet_name
        self.pretty_headers = pretty_headers
        self.long_format = long_format
        self.with_table = with_table
        self.autosize = autosize
        self.save_checkpoint_s = save_checkpoint_s

        self._stop = threading.Event()
        self._wb = None
        self._path = None
        self._header = []      # orden de columnas para 'rt'
        self._maxlen = {}      # autosize
        self._last_save = 0.0
        self.rows_written = 0  # expuesto para UI
        self.drops = 0         # (por si lo quieres usar)

    def stop(self):
        self._stop.set()

    def _ensure_workbook(self):
        date = datetime.now().strftime("%Y-%m-%d")
        base = self.path_template.format(date=date)
        base = os.path.normpath(base)
        os.makedirs(os.path.dirname(base), exist_ok=True)

        if self._wb is None:
            # SIEMPRE crea un libro nuevo con índice incremental
            self._path = _next_indexed_path(base)
            self._wb = Workbook()

            # 1) Hoja RT primero
            ws_rt = self._wb.active
            ws_rt.title = self.sheet_name
            self._header = ["timestamp"]
            ws_rt.append(self._header)

            # 2) Hoja RAW segundo
            ws_raw = self._wb.create_sheet("raw")
            ws_raw.append(["timestamp", "json"])

            self._last_save = 0.0
            self.rows_written = 0
            self._maxlen = {1: len("timestamp")}

    def _checkpoint_save(self):
        if (time.time() - self._last_save) >= self.save_checkpoint_s:
            # ajusta tabla y ancho columnas antes de guardar
            ws_rt = self._wb[self.sheet_name]
            last_col = get_column_letter(len(self._header))
            last_row = ws_rt.max_row
            if self.with_table:
                name = "tbl_rt"
                if name in ws_rt.tables:
                    tbl = ws_rt.tables[name]
                    tbl.ref = f"A1:{last_col}{last_row}"
                else:
                    tbl = Table(displayName=name, ref=f"A1:{last_col}{last_row}")
                    style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
                    tbl.tableStyleInfo = style
                    ws_rt.add_table(tbl)
            if self.autosize:
                for idx in range(1, len(self._header) + 1):
                    col_letter = get_column_letter(idx)
                    ws_rt.column_dimensions[col_letter].width = max(10, min(50, self._maxlen.get(idx, 12)))

            self._wb.save(self._path)
            self._last_save = time.time()

    def run(self):
        batch = []
        last = time.time()
        while not self._stop.is_set():
            try:
                item = self.q.get(timeout=0.2)
                batch.append(item)
                if len(batch) >= self.flush_every or (time.time() - last) >= self.flush_interval:
                    self._flush(batch)
                    batch.clear()
                    last = time.time()
                    self._checkpoint_save()
            except queue.Empty:
                # nada en cola; si hay batch pendiente por tiempo, flush
                if batch and (time.time() - last) >= self.flush_interval:
                    self._flush(batch)
                    batch.clear()
                    last = time.time()
                    self._checkpoint_save()

        # salida limpia
        if batch:
            self._flush(batch)
        self._checkpoint_save()

    # --- internos ---
    def _flush(self, batch):
        self._ensure_workbook()
        ws_rt = self._wb[self.sheet_name]
        ws_raw = self._wb["raw"]

        header_changed = False

        for sample in batch:
            # 1) hoja RAW
            ts = sample.get("timestamp", time.time())
            ws_raw.append([ts, json.dumps(sample, ensure_ascii=False, separators=(",", ":"))])

            # 2) hoja RT (wide)
            flat = _flatten(sample)
            # agrega columnas nuevas si aparecen
            for key in flat.keys():
                if key not in self._header:
                    self._header.append(key)
                    header_changed = True

            if header_changed:
                # reescribe header (fila 1), bonito si corresponde
                for idx, key in enumerate(self._header, start=1):
                    text = _beautify(key) if (self.pretty_headers and key != "timestamp") else key
                    ws_rt.cell(row=1, column=idx, value=text)
                    self._maxlen[idx] = max(self._maxlen.get(idx, 0), len(str(text)))
                header_changed = False

            # arma fila
            row = []
            for idx, col in enumerate(self._header, start=1):
                if col == "timestamp":
                    # escribe fecha/hora legible con milisegundo
                    dt = datetime.fromtimestamp(ts)
                    text = dt.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
                    row.append(text)
                    self._maxlen[idx] = max(self._maxlen.get(idx, 0), len(text))
                else:
                    val = flat.get(col, None)
                    row.append(val)
                    self._maxlen[idx] = max(self._maxlen.get(idx, 0), len(str(val)))

            ws_rt.append(row)
            self.rows_written += 1
