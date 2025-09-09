import os, asyncio, json, time
from datetime import datetime, timezone
from typing import Dict, Any, List, Tuple
from threading import Lock

from opcua import Client, ua
import aiosqlite
import pandas as pd
from fastapi import FastAPI, WebSocket

# ===== OPC UA =====
OPCUA_URL     = "opc.tcp://192.168.18.6:4840"
OPCUA_USER    = "boschrexroth"
OPCUA_PASS    = "boschrexroth"
USE_SECURITY  = False  # True si luego configuras certificados y endpoint seguro

# Folder base en Data Layer (cópialo de la UI)
START_NODE_ID = "ns=2;s=plc/app/Application/sym/PLC_PRG"

# ===== Tiempos / buffers =====
SAMPLING_MS   = 0          # 0 = lo más rápido posible (si el server lo permite)
PUBLISHING_MS = 50
QUEUE_SIZE    = 20000
BATCH_FLUSH   = 200
EXCEL_PERIOD  = 5          # seg entre volcados a Excel
WS_SEND_MS    = 100        # periodo de envío por WebSocket (ms)
CHUNK_CREATE  = 500        # crea MonitoredItems por tandas

# ===== Storage =====
os.makedirs("logs", exist_ok=True)
DB_PATH = "logs/opc_live.db"

# ===== Estado en memoria + locks =====
last_snapshot: Dict[str, Any] = {}
insert_queue: List[Tuple[str, str, str, int]] = []
alias_map: Dict[str, str] = {}  # nodeid -> alias (BrowseName)
LOCK = Lock()

# ---------------- SQLite ----------------

async def ensure_db():
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("PRAGMA journal_mode=WAL;")
        await db.execute("PRAGMA synchronous=FULL;")
        await db.execute("""
        CREATE TABLE IF NOT EXISTS samples(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ts TEXT NOT NULL,
            node TEXT NOT NULL,
            value TEXT,
            status INTEGER NOT NULL
        );
        """)
        await db.commit()

async def writer_task():
    await ensure_db()
    last = time.time()
    while True:
        batch = None
        with LOCK:
            if len(insert_queue) >= BATCH_FLUSH or (insert_queue and time.time()-last > 0.5):
                batch, insert_queue[:] = insert_queue[:], []
        if batch:
            async with aiosqlite.connect(DB_PATH) as db:
                await db.executemany(
                    "INSERT INTO samples(ts,node,value,status) VALUES (?,?,?,?)",
                    batch
                )
                await db.commit()
            last = time.time()
        await asyncio.sleep(0.05)

# ---------------- Excel incremental (encabezados por alias) ----------------

def _daily_xlsx_path() -> str:
    return f"logs/OPC_{datetime.now().strftime('%Y-%m-%d')}.xlsx"

def _ordered_aliases() -> List[str]:
    with LOCK:
        # columnas ordenadas: por alias (BrowseName) estable; si no hay alias, usa nodeid
        pairs = [(alias_map.get(nid, nid), nid) for nid in alias_map.keys()]
    pairs.sort(key=lambda x: x[0].lower())
    return [p[0] for p in pairs]  # solo alias ordenados

async def excel_exporter_task():
    """
    Toma filas nuevas de SQLite y las vuelca a un Excel “ancho”:
    columnas = ['ts'] + alias1 + alias2 + ...
    cada ejecución agrega filas debajo con los valores más recientes por ts.
    """
    last_id = 0
    wrote_header_for_today = False
    last_day = datetime.now().day

    while True:
        try:
            # Detecta cambio de día (nuevo archivo y nuevo header)
            today = datetime.now().day
            if today != last_day:
                wrote_header_for_today = False
                last_day = today

            # lee incremento
            import sqlite3
            con = sqlite3.connect(DB_PATH)
            df = pd.read_sql_query(
                f"SELECT * FROM samples WHERE id>{last_id} ORDER BY id",
                con
            )
            con.close()

            if not df.empty:
                last_id = int(df["id"].iloc[-1])

                # mapea node -> alias para columnas “bonitas”
                with LOCK:
                    # crea columna 'alias' usando alias_map; fallback al nodeid si no existe
                    df["alias"] = df["node"].map(lambda nid: alias_map.get(nid, nid))

                # pivot por alias para tener 1 fila por ts y columnas por variable
                piv = df.pivot_table(index="ts", columns="alias", values="value",
                                     aggfunc="last").reset_index()

                # orden de columnas fijo
                cols = ["ts"] + _ordered_aliases()
                # incluye columnas nuevas si aparecieron
                for c in piv.columns:
                    if c not in cols:
                        cols.append(c)
                piv = piv.reindex(columns=cols, fill_value=None)

                xlsx = _daily_xlsx_path()
                if not os.path.exists(xlsx):
                    # primer archivo del día -> escribe encabezado
                    with pd.ExcelWriter(xlsx, engine="openpyxl") as xw:
                        piv.to_excel(xw, sheet_name="data", index=False)
                    wrote_header_for_today = True
                else:
                    # agrega sin repetir encabezado
                    from openpyxl import load_workbook
                    wb = load_workbook(xlsx)
                    if "data" not in wb.sheetnames:
                        with pd.ExcelWriter(xlsx, engine="openpyxl", mode="a") as xw:
                            piv.to_excel(xw, sheet_name="data", index=False)
                        wrote_header_for_today = True
                    else:
                        ws = wb["data"]
                        start_row = ws.max_row + (0 if ws.max_row == 1 else 1)
                        with pd.ExcelWriter(xlsx, engine="openpyxl", mode="a", if_sheet_exists="overlay") as xw:
                            piv.to_excel(
                                xw, sheet_name="data", index=False,
                                header=not wrote_header_for_today,
                                startrow=ws.max_row
                            )
                        wrote_header_for_today = True
        except Exception as e:
            print("Exporter warn:", e)

        await asyncio.sleep(EXCEL_PERIOD)

# ---------------- OPC UA ----------------

def browse_all_variables(session: Client, start_nodeid_str: str):
    root = session.get_node(ua.NodeId.from_string(start_nodeid_str))
    stack = [root]
    vars_ = []
    while stack:
        n = stack.pop()
        try:
            children = n.get_children()
        except Exception:
            continue
        for c in children:
            try:
                nc = c.get_node_class()
            except Exception:
                continue
            if nc == ua.NodeClass.Variable:
                vars_.append(c)
            elif nc in (ua.NodeClass.Object, ua.NodeClass.View):
                stack.append(c)
    return vars_

def make_readvalueid(nodeid: ua.NodeId) -> ua.ReadValueId:
    rid = ua.ReadValueId()
    rid.NodeId = nodeid
    rid.AttributeId = ua.AttributeIds.Value
    return rid

def status_to_int(sc) -> int:
    if sc is None:
        return 0
    try:
        return int(sc)
    except Exception:
        return int(getattr(sc, "value", 0))

class SubHandler:
    def datachange_notification(self, node, val, data):
        # timestamp
        src_ts = getattr(getattr(data, "monitored_item", None), "Value", None)
        src_ts = getattr(src_ts, "SourceTimestamp", None)
        ts = (src_ts or datetime.now(timezone.utc)).astimezone().isoformat()

        nodeid = node.nodeid.to_string()
        sc_obj = getattr(getattr(data, "monitored_item", None), "Value", None)
        sc_obj = getattr(sc_obj, "StatusCode", None)
        status = status_to_int(sc_obj)
        sval = str(val)

        try:
            alias = node.get_browse_name().Name
        except Exception:
            alias = nodeid

        with LOCK:
            # guarda alias para excel/orden
            alias_map[nodeid] = alias
            # persiste
            insert_queue.append((ts, nodeid, sval, status))
            # último snapshot para WS
            last_snapshot[alias] = val
            last_snapshot["_timestamp"] = ts

async def opc_task():
    while True:
        client = None
        try:
            client = Client(OPCUA_URL, timeout=4)
            client.set_user(OPCUA_USER)
            client.set_password(OPCUA_PASS)

            if USE_SECURITY:
                # configurar si activas endpoint seguro
                pass

            client.connect()
            variables = browse_all_variables(client, START_NODE_ID)
            print(f"[OPC] Variables encontradas: {len(variables)}")

            # llena alias_map desde el browse para encabezados desde inicio
            with LOCK:
                for v in variables:
                    try:
                        alias_map[v.nodeid.to_string()] = v.get_browse_name().Name
                    except Exception:
                        alias_map[v.nodeid.to_string()] = v.nodeid.to_string()

            sub = client.create_subscription(PUBLISHING_MS, SubHandler())

            # Prepara requests (sin Filter para máxima compatibilidad)
            reqs = []
            for i, varnode in enumerate(variables, start=1):
                params = ua.MonitoringParameters()
                params.ClientHandle     = i
                params.SamplingInterval = SAMPLING_MS
                params.QueueSize        = QUEUE_SIZE
                params.DiscardOldest    = True

                mi = ua.MonitoredItemCreateRequest()
                mi.ItemToMonitor       = make_readvalueid(varnode.nodeid)
                mi.MonitoringMode      = ua.MonitoringMode.Reporting
                mi.RequestedParameters = params

                reqs.append(mi)

            # Crea en tandas (firma vieja: solo la lista)
            for k in range(0, len(reqs), CHUNK_CREATE):
                chunk = reqs[k:k+CHUNK_CREATE]
                results = sub.create_monitored_items(chunk)
                codes = []
                for r in results:
                    sc = getattr(r, "StatusCode", r)
                    try: codes.append(int(sc))
                    except Exception: codes.append(int(getattr(sc, "value", 0)))
                ok = sum(1 for c in codes if c == 0)
                print(f"[OPC] Creó {ok}/{len(chunk)} items en el bloque {k//CHUNK_CREATE+1}  StatusCodes={codes}")

            # keep-alive
            while True:
                await asyncio.sleep(0.2)

        except Exception as e:
            print("OPC reconnect in 2s ->", e)
            await asyncio.sleep(2)
        finally:
            try:
                if client: client.disconnect()
            except: pass

# ---------------- FastAPI (WebSocket + health + startup) ----------------

app = FastAPI()
_bg_tasks: List[asyncio.Task] = []

@app.on_event("startup")
async def _startup():
    # arranca todo con uvicorn
    await ensure_db()
    _bg_tasks[:] = [
        asyncio.create_task(opc_task(), name="opc_task"),
        asyncio.create_task(writer_task(), name="writer_task"),
        asyncio.create_task(excel_exporter_task(), name="excel_exporter_task"),
    ]
    print("[APP] background tasks started")

@app.on_event("shutdown")
async def _shutdown():
    for t in _bg_tasks:
        t.cancel()
    await asyncio.gather(*_bg_tasks, return_exceptions=True)
    print("[APP] background tasks stopped")

@app.get("/health")
def health():
    with LOCK:
        sz = len(last_snapshot)
        cols = _ordered_aliases()
    return {"opc": "ok", "vars_cache": sz, "columns": ["ts"] + cols}

@app.websocket("/ws")
async def ws_endpoint(ws: WebSocket):
    await ws.accept()
    try:
        while True:
            payload = None
            with LOCK:
                if last_snapshot:
                    payload = json.dumps([last_snapshot], default=str)
            if payload:
                await ws.send_text(payload)
            await asyncio.sleep(WS_SEND_MS / 1000)
    except Exception as e:
        print("WS client off:", e)

# (sin main; se arranca con uvicorn)
